#!/usr/bin/env python3
"""
Validation Report Generator for DOCX to XML Pipeline

Generates Excel (XLSX) reports documenting XML validation errors against
the RittDoc DTD specification. Matches the output format of the PDFtoXML
pipeline's validation report.

Usage:
    python validation_report.py output/document_docbook42.xml
    python validation_report.py output/document_rittdoc.zip
"""

from __future__ import annotations

import os
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union
from io import BytesIO
import zipfile

from lxml import etree

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class ValidationError:
    """Represents a single validation error."""
    xml_file: str = ""
    line_number: int = 0
    column_number: int = 0
    error_type: str = ""
    error_description: str = ""
    severity: str = "Error"  # Error, Warning, Info

    def to_dict(self) -> Dict[str, Any]:
        return {
            "xml_file": self.xml_file,
            "line_number": self.line_number,
            "column_number": self.column_number,
            "error_type": self.error_type,
            "error_description": self.error_description,
            "severity": self.severity,
        }


@dataclass
class VerificationItem:
    """Represents an item needing manual verification."""
    xml_file: str = ""
    line_number: int = 0
    fix_type: str = ""
    fix_description: str = ""
    verification_reason: str = ""
    suggestion: str = ""


@dataclass
class ValidationResult:
    """Complete validation result."""
    errors: List[ValidationError] = field(default_factory=list)
    warnings: List[ValidationError] = field(default_factory=list)
    info: List[ValidationError] = field(default_factory=list)
    verifications: List[VerificationItem] = field(default_factory=list)
    xml_file: str = ""
    is_valid: bool = False
    total_errors: int = 0
    total_warnings: int = 0


# ============================================================================
# RITTDOC DTD RULES
# ============================================================================

# Elements excluded from RittDoc DTD
EXCLUDED_ELEMENTS = {
    'informalfigure', 'informaltable', 'variablelist', 'simplelist',
    'example', 'procedure', 'div', 'span', 'b', 'i', 'u',
    'bridgehead', 'abstract', 'sidebar', 'tip', 'note', 'warning',
    'caution', 'important',
}

# Elements that require a <title> child
TITLE_REQUIRED = {'chapter', 'sect1', 'sect2', 'sect3', 'sect4', 'sect5',
                  'figure', 'table', 'book'}

# Elements that require specific attributes
REQUIRED_ATTRIBUTES = {
    'tgroup': ['cols'],
    'imagedata': ['fileref'],
    'chapter': ['id'],
    'sect1': ['id'],
    'sect2': ['id'],
    'sect3': ['id'],
    'figure': ['id'],
    'table': ['id'],
}

# ID format patterns
ID_PATTERNS = {
    'chapter': re.compile(r'^ch\d{4}$'),
    'sect1': re.compile(r'^ch\d{4}s\d{4}$'),
    'sect2': re.compile(r'^ch\d{4}s\d{4}$'),
    'sect3': re.compile(r'^ch\d{4}s\d{4}$'),
    'figure': re.compile(r'^ch\d{4}s\d{4}fg\d{2}$'),
    'table': re.compile(r'^ch\d{4}s\d{4}tb\d{2}$'),
}

# Valid figure structure: figure -> title + mediaobject -> imageobject -> imagedata
FIGURE_REQUIRED_CHILDREN = ['title', 'mediaobject']

# Valid table structure: table -> title + tgroup -> (colspec*, thead?, tbody)
TABLE_REQUIRED_CHILDREN = ['title', 'tgroup']


# ============================================================================
# VALIDATOR CLASS
# ============================================================================

class RittDocValidator:
    """
    Validates DocBook XML against RittDoc DTD rules.

    Checks for:
    - Excluded elements
    - Required titles
    - Required attributes
    - ID format compliance
    - Figure structure
    - Table structure (CALS format)
    - Section hierarchy
    - Content model compliance
    """

    def __init__(self):
        self.errors: List[ValidationError] = []
        self.verifications: List[VerificationItem] = []
        self._seen_ids: set = set()

    def validate(self, xml_content: str, xml_file: str = "Book.xml") -> ValidationResult:
        """
        Validate XML content against RittDoc DTD rules.

        Args:
            xml_content: XML string to validate
            xml_file: Filename for error reporting

        Returns:
            ValidationResult with all errors and warnings
        """
        self.errors = []
        self.verifications = []
        self._seen_ids = set()

        result = ValidationResult(xml_file=xml_file)

        # Parse XML
        try:
            parser = etree.XMLParser(recover=True, resolve_entities=False)
            root = etree.fromstring(xml_content.encode('utf-8'), parser=parser)
        except Exception as e:
            result.errors.append(ValidationError(
                xml_file=xml_file,
                error_type="XML Parse Error",
                error_description=f"Failed to parse XML: {str(e)}",
                severity="Error"
            ))
            result.total_errors = 1
            return result

        # Run all validation checks
        self._validate_element(root, xml_file)

        # Check DOCTYPE
        self._check_doctype(xml_content, xml_file)

        # Categorize results
        for err in self.errors:
            if err.severity == "Error":
                result.errors.append(err)
            elif err.severity == "Warning":
                result.warnings.append(err)
            else:
                result.info.append(err)

        result.verifications = self.verifications
        result.total_errors = len(result.errors)
        result.total_warnings = len(result.warnings)
        result.is_valid = result.total_errors == 0

        return result

    def _validate_element(self, elem, xml_file: str, depth: int = 0):
        """Recursively validate an element and its children."""
        tag = self._local_name(elem.tag)

        # Get source line if available
        line = elem.sourceline or 0

        # Check for excluded elements
        if tag in EXCLUDED_ELEMENTS:
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=line,
                error_type="Excluded Element",
                error_description=f"Element <{tag}> is not allowed in RittDoc DTD. "
                                  f"Use an appropriate alternative.",
                severity="Error"
            ))

        # Check required titles
        if tag in TITLE_REQUIRED:
            title_elem = elem.find('title')
            # For <book>, title can be inside <bookinfo>
            if title_elem is None and tag == 'book':
                bookinfo = elem.find('bookinfo')
                if bookinfo is not None:
                    title_elem = bookinfo.find('title')
            if title_elem is None:
                self.errors.append(ValidationError(
                    xml_file=xml_file,
                    line_number=line,
                    error_type="Missing Title",
                    error_description=f"Element <{tag}> requires a <title> child element.",
                    severity="Error"
                ))

        # Check required attributes
        if tag in REQUIRED_ATTRIBUTES:
            for attr in REQUIRED_ATTRIBUTES[tag]:
                if elem.get(attr) is None:
                    self.errors.append(ValidationError(
                        xml_file=xml_file,
                        line_number=line,
                        error_type="Missing Attribute",
                        error_description=f"Element <{tag}> requires attribute '{attr}'.",
                        severity="Error"
                    ))

        # Check ID format
        if tag in ID_PATTERNS:
            elem_id = elem.get('id', '')
            if elem_id:
                if elem_id in self._seen_ids:
                    self.errors.append(ValidationError(
                        xml_file=xml_file,
                        line_number=line,
                        error_type="Duplicate ID",
                        error_description=f"Duplicate id='{elem_id}' found on <{tag}>.",
                        severity="Error"
                    ))
                self._seen_ids.add(elem_id)

                if not ID_PATTERNS[tag].match(elem_id):
                    self.errors.append(ValidationError(
                        xml_file=xml_file,
                        line_number=line,
                        error_type="Invalid ID Format",
                        error_description=f"ID '{elem_id}' on <{tag}> does not match "
                                          f"expected pattern. Expected: {ID_PATTERNS[tag].pattern}",
                        severity="Warning"
                    ))

        # Check figure structure
        if tag == 'figure':
            self._validate_figure(elem, xml_file, line)

        # Check table structure
        if tag == 'table':
            self._validate_table(elem, xml_file, line)

        # Check section hierarchy
        if tag in ('sect2', 'sect3'):
            self._validate_section_hierarchy(elem, tag, xml_file, line)

        # Check imagedata attributes
        if tag == 'imagedata':
            self._validate_imagedata(elem, xml_file, line)

        # Recurse into children
        for child in elem:
            self._validate_element(child, xml_file, depth + 1)

    def _validate_figure(self, elem, xml_file: str, line: int):
        """Validate figure element structure."""
        children = [self._local_name(c.tag) for c in elem]

        if 'mediaobject' not in children:
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=line,
                error_type="Invalid Figure Structure",
                error_description="<figure> must contain a <mediaobject> element.",
                severity="Error"
            ))
        else:
            # Check mediaobject contains imageobject
            mediaobj = elem.find('mediaobject')
            if mediaobj is not None:
                imgobj = mediaobj.find('imageobject')
                if imgobj is None:
                    self.errors.append(ValidationError(
                        xml_file=xml_file,
                        line_number=line,
                        error_type="Invalid Figure Structure",
                        error_description="<mediaobject> must contain an <imageobject> element.",
                        severity="Error"
                    ))
                else:
                    imgdata = imgobj.find('imagedata')
                    if imgdata is None:
                        self.errors.append(ValidationError(
                            xml_file=xml_file,
                            line_number=line,
                            error_type="Invalid Figure Structure",
                            error_description="<imageobject> must contain an <imagedata> element.",
                            severity="Error"
                        ))

    def _validate_table(self, elem, xml_file: str, line: int):
        """Validate table element structure (CALS format)."""
        tgroup = elem.find('tgroup')
        if tgroup is None:
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=line,
                error_type="Invalid Table Structure",
                error_description="<table> must contain a <tgroup> element (CALS format).",
                severity="Error"
            ))
            return

        # Check cols attribute
        cols = tgroup.get('cols')
        if cols is None:
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=tgroup.sourceline or line,
                error_type="Missing Attribute",
                error_description="<tgroup> requires 'cols' attribute with column count.",
                severity="Error"
            ))
        elif not cols.isdigit() or int(cols) < 1:
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=tgroup.sourceline or line,
                error_type="Invalid Attribute Value",
                error_description=f"<tgroup cols='{cols}'> must be a positive integer.",
                severity="Error"
            ))

        # Check for HTML table elements (should be CALS)
        html_elements = elem.findall('.//{http://www.w3.org/1999/xhtml}tr') or \
                        elem.findall('.//tr') or elem.findall('.//td') or elem.findall('.//th')
        if html_elements:
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=line,
                error_type="Invalid Table Format",
                error_description="Table contains HTML elements (tr/td/th). Must use CALS format "
                                  "(tgroup/row/entry).",
                severity="Error"
            ))

        # Check tbody exists
        tbody = tgroup.find('tbody')
        if tbody is None:
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=tgroup.sourceline or line,
                error_type="Invalid Table Structure",
                error_description="<tgroup> should contain a <tbody> element.",
                severity="Warning"
            ))

    def _validate_section_hierarchy(self, elem, tag: str, xml_file: str, line: int):
        """Validate section nesting is correct."""
        parent = elem.getparent()
        if parent is None:
            return

        parent_tag = self._local_name(parent.tag)

        if tag == 'sect2' and parent_tag != 'sect1':
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=line,
                error_type="Invalid Section Hierarchy",
                error_description=f"<sect2> must be a child of <sect1>, found inside <{parent_tag}>.",
                severity="Error"
            ))

        if tag == 'sect3' and parent_tag != 'sect2':
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=line,
                error_type="Invalid Section Hierarchy",
                error_description=f"<sect3> must be a child of <sect2>, found inside <{parent_tag}>.",
                severity="Error"
            ))

    def _validate_imagedata(self, elem, xml_file: str, line: int):
        """Validate imagedata attributes."""
        fileref = elem.get('fileref', '')
        if not fileref:
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=line,
                error_type="Missing Attribute",
                error_description="<imagedata> requires 'fileref' attribute.",
                severity="Error"
            ))
        else:
            # Check width and scalefit
            if not elem.get('width'):
                self.errors.append(ValidationError(
                    xml_file=xml_file,
                    line_number=line,
                    error_type="Missing Attribute",
                    error_description="<imagedata> should have 'width' attribute.",
                    severity="Warning"
                ))
            if not elem.get('scalefit'):
                self.errors.append(ValidationError(
                    xml_file=xml_file,
                    line_number=line,
                    error_type="Missing Attribute",
                    error_description="<imagedata> should have 'scalefit' attribute.",
                    severity="Warning"
                ))

            # Verify multimedia path format
            if 'multimedia/' in fileref:
                filename = fileref.replace('multimedia/', '')
                if not re.match(r'Ch\d{4}s\d{4}fg\d{2}\.\w+', filename):
                    self.verifications.append(VerificationItem(
                        xml_file=xml_file,
                        line_number=line,
                        fix_type="Figure Naming",
                        fix_description=f"Image filename '{filename}' may not follow convention.",
                        verification_reason="Expected format: Ch0000s0000fg00.ext",
                        suggestion=f"Rename to follow Ch{{chap}}s{{sect}}fg{{num}}.ext pattern"
                    ))

    def _check_doctype(self, xml_content: str, xml_file: str):
        """Check for proper DOCTYPE declaration."""
        if '<!DOCTYPE' not in xml_content:
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=1,
                error_type="Missing DOCTYPE",
                error_description="XML is missing DOCTYPE declaration. "
                                  "Expected: <!DOCTYPE book PUBLIC \"-//RIS Dev//DTD DocBook V4.3 "
                                  "-Based Variant V1.1//EN\" ...>",
                severity="Warning"
            ))
        elif 'RittDocBook.dtd' not in xml_content and 'RIS Dev' not in xml_content:
            self.errors.append(ValidationError(
                xml_file=xml_file,
                line_number=1,
                error_type="Invalid DOCTYPE",
                error_description="DOCTYPE does not reference RittDoc DTD. "
                                  "Expected PUBLIC \"-//RIS Dev//DTD DocBook V4.3 -Based Variant V1.1//EN\"",
                severity="Warning"
            ))

    def _local_name(self, tag: str) -> str:
        """Extract local name from potentially namespaced tag."""
        if '}' in tag:
            return tag.split('}')[1]
        return tag


# ============================================================================
# REPORT GENERATOR
# ============================================================================

class ValidationReportGenerator:
    """
    Generates XLSX validation reports.

    Report structure:
    - Sheet 1: Validation Errors (Error #, Severity, XML File, Line, Column, Type, Description)
    - Sheet 2: Summary (statistics by severity and type)
    - Sheet 3: Manual Verification (items needing content review)
    """

    # Styling constants
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    INFO_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def generate_report(
        self,
        result: ValidationResult,
        output_path: Union[str, Path],
        title: str = "Validation Report"
    ) -> str:
        """
        Generate an XLSX validation report.

        Args:
            result: ValidationResult from validation
            output_path: Path for the output XLSX file
            title: Report title

        Returns:
            Path to the generated report
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for report generation. "
                              "Install with: pip install openpyxl")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()

        # Sheet 1: Validation Errors
        self._create_errors_sheet(wb, result)

        # Sheet 2: Summary
        self._create_summary_sheet(wb, result, title)

        # Sheet 3: Manual Verification
        self._create_verification_sheet(wb, result)

        # Save with retry logic
        self._save_workbook_with_retry(wb, output_path)

        return str(output_path)

    def _create_errors_sheet(self, wb: openpyxl.Workbook, result: ValidationResult):
        """Create the Validation Errors sheet."""
        ws = wb.active
        ws.title = "Validation Errors"

        # Headers
        headers = ["Error #", "Severity", "XML File", "Line Number",
                   "Column", "Error Type", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER

        # Combine all errors sorted by severity
        all_errors = []
        all_errors.extend(result.errors)
        all_errors.extend(result.warnings)
        all_errors.extend(result.info)

        # Data rows
        for i, err in enumerate(all_errors, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).border = self.BORDER
            severity_cell = ws.cell(row=row, column=2, value=err.severity)
            severity_cell.border = self.BORDER
            ws.cell(row=row, column=3, value=err.xml_file).border = self.BORDER
            ws.cell(row=row, column=4, value=err.line_number or "N/A").border = self.BORDER
            ws.cell(row=row, column=5, value=err.column_number or "N/A").border = self.BORDER
            ws.cell(row=row, column=6, value=err.error_type).border = self.BORDER
            desc_cell = ws.cell(row=row, column=7, value=err.error_description)
            desc_cell.border = self.BORDER
            desc_cell.alignment = Alignment(wrap_text=True)

            # Color-code severity
            if err.severity == "Error":
                severity_cell.fill = self.ERROR_FILL
            elif err.severity == "Warning":
                severity_cell.fill = self.WARNING_FILL
            else:
                severity_cell.fill = self.INFO_FILL

        # Column widths
        col_widths = [8, 10, 20, 12, 8, 22, 60]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header
        ws.freeze_panes = "A2"

    def _create_summary_sheet(self, wb: openpyxl.Workbook, result: ValidationResult,
                              title: str):
        """Create the Summary sheet."""
        ws = wb.create_sheet("Summary")

        # Title
        ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=3, column=1, value=f"XML File: {result.xml_file}")

        # Overall status
        row = 5
        status = "PASSED" if result.is_valid else "FAILED"
        status_cell = ws.cell(row=row, column=1, value=f"Validation Status: {status}")
        status_cell.font = Font(size=12, bold=True,
                                color="008000" if result.is_valid else "FF0000")

        # Statistics by severity
        row = 7
        ws.cell(row=row, column=1, value="Statistics by Severity").font = Font(bold=True)
        row += 1
        headers = ["Severity", "Count"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER

        severity_counts = {
            "Error": len(result.errors),
            "Warning": len(result.warnings),
            "Info": len(result.info),
        }

        for severity, count in severity_counts.items():
            row += 1
            ws.cell(row=row, column=1, value=severity).border = self.BORDER
            ws.cell(row=row, column=2, value=count).border = self.BORDER

        row += 1
        ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=row, column=1).border = self.BORDER
        ws.cell(row=row, column=2, value=sum(severity_counts.values())).border = self.BORDER
        ws.cell(row=row, column=2).font = Font(bold=True)

        # Statistics by error type
        row += 2
        ws.cell(row=row, column=1, value="Statistics by Error Type").font = Font(bold=True)
        row += 1
        headers = ["Error Type", "Count"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER

        all_errors = result.errors + result.warnings + result.info
        type_counts: Dict[str, int] = {}
        for err in all_errors:
            type_counts[err.error_type] = type_counts.get(err.error_type, 0) + 1

        for error_type, count in sorted(type_counts.items(), key=lambda x: -x[1]):
            row += 1
            ws.cell(row=row, column=1, value=error_type).border = self.BORDER
            ws.cell(row=row, column=2, value=count).border = self.BORDER

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12

    def _create_verification_sheet(self, wb: openpyxl.Workbook, result: ValidationResult):
        """Create the Manual Verification sheet."""
        ws = wb.create_sheet("Manual Verification")

        # Headers
        headers = ["#", "XML File", "Line", "Fix Type",
                   "Description", "Reason", "Suggestion"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER

        # Data rows
        for i, item in enumerate(result.verifications, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i).border = self.BORDER
            ws.cell(row=row, column=2, value=item.xml_file).border = self.BORDER
            ws.cell(row=row, column=3, value=item.line_number or "N/A").border = self.BORDER
            ws.cell(row=row, column=4, value=item.fix_type).border = self.BORDER
            ws.cell(row=row, column=5, value=item.fix_description).border = self.BORDER
            ws.cell(row=row, column=6, value=item.verification_reason).border = self.BORDER
            ws.cell(row=row, column=7, value=item.suggestion).border = self.BORDER

        # Column widths
        col_widths = [6, 18, 8, 18, 40, 30, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze header
        ws.freeze_panes = "A2"

    def _save_workbook_with_retry(self, wb: openpyxl.Workbook, path: Path,
                                   max_retries: int = 3):
        """Save workbook with retry logic for file-locking issues."""
        for attempt in range(max_retries):
            try:
                wb.save(str(path))
                return
            except PermissionError:
                if attempt < max_retries - 1:
                    wait = (attempt + 1) * 2
                    print(f"  File locked, retrying in {wait}s...")
                    time.sleep(wait)
                else:
                    raise


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def validate_xml(xml_content: str, xml_file: str = "Book.xml") -> ValidationResult:
    """Validate XML content against RittDoc DTD rules."""
    validator = RittDocValidator()
    return validator.validate(xml_content, xml_file)


def generate_validation_report(
    xml_content: str,
    output_path: Union[str, Path],
    xml_file: str = "Book.xml",
    title: str = "RittDoc DTD Validation Report"
) -> str:
    """
    Validate XML and generate an XLSX report.

    Args:
        xml_content: XML string to validate
        output_path: Path for the output XLSX file
        xml_file: Filename for error reporting
        title: Report title

    Returns:
        Path to the generated report
    """
    validator = RittDocValidator()
    result = validator.validate(xml_content, xml_file)

    generator = ValidationReportGenerator()
    report_path = generator.generate_report(result, output_path, title)

    return report_path


def validate_package(package_path: Union[str, Path], output_dir: Optional[Path] = None) -> str:
    """
    Validate a RittDoc ZIP package and generate report.

    Args:
        package_path: Path to the ZIP package
        output_dir: Directory for the report (defaults to same as package)

    Returns:
        Path to the generated report
    """
    package_path = Path(package_path)

    if not package_path.exists():
        raise FileNotFoundError(f"Package not found: {package_path}")

    # Extract XML from package
    xml_content = ""
    xml_file = "Book.xml"

    if package_path.suffix.lower() == '.zip':
        with zipfile.ZipFile(package_path, 'r') as zf:
            # Find Book.xml
            for name in zf.namelist():
                if name.endswith('.xml') and 'metadata' not in name.lower():
                    xml_content = zf.read(name).decode('utf-8')
                    xml_file = name
                    break
    else:
        # Assume it's an XML file directly
        with open(package_path, 'r', encoding='utf-8') as f:
            xml_content = f.read()
        xml_file = package_path.name

    if not xml_content:
        raise ValueError("No XML content found in package")

    # Output path
    if output_dir is None:
        output_dir = package_path.parent

    stem = package_path.stem.replace('_rittdoc', '').replace('_docbook42', '')
    report_path = output_dir / f"{stem}_validation_report.xlsx"

    return generate_validation_report(xml_content, report_path, xml_file,
                                      f"Validation Report: {stem}")


# ============================================================================
# CLI
# ============================================================================

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python validation_report.py <xml_file_or_zip> [output_dir]")
        print("\nExamples:")
        print("  python validation_report.py output/document_rittdoc.zip")
        print("  python validation_report.py output/document_docbook42.xml")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        sys.exit(1)

    print(f"Validating: {input_path}")
    try:
        report_path = validate_package(input_path, output_dir)
        print(f"Validation report generated: {report_path}")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)
